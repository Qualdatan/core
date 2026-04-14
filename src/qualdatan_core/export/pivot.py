"""Pivot-Tabellen-Export: Wide-Format-Excel mit allen Codings.

Erzeugt eine einzelne Excel-Datei (``pivot_results.xlsx``) mit dem Sheet
``Codierungen``: pro Zeile genau ein kodiertes Segment (Interview ODER PDF),
Spalten = Dimensionen fuer Pivot-Tabellen / Charts.

Defensiv: Wenn keine Interview-Analyse vorliegt, werden nur PDF-Codings
geschrieben (und umgekehrt). Bei 0 Codings insgesamt wird die Datei NICHT
erzeugt und 0 zurueckgegeben.
"""

from __future__ import annotations

from collections.abc import Iterable
from pathlib import Path

from ..models import AnalysisResult
from ..run_context import RunContext

# ---------------------------------------------------------------------------
# Spalten-Definition
# ---------------------------------------------------------------------------

COLUMNS = [
    "Run",
    "Quelle",
    "Company",
    "Projekt",
    "Datei",
    "Pfad",
    "Seite",
    "Code",
    "Hauptkategorie",
    "Code-Name",
    "Text",
    "Begründung",
    "Zeichen-Start",
    "Zeichen-Ende",
    "Confidence",
]

MAX_TEXT_LEN = 500


def _truncate(text: str, max_len: int = MAX_TEXT_LEN) -> str:
    text = (text or "").replace("\r", " ").strip()
    if len(text) <= max_len:
        return text
    return text[: max_len - 4].rstrip() + " ..."


def _hauptkategorie_from_code(code_id: str) -> str:
    """Erste sinnvolle Komponente einer Code-ID (z.B. 'B-03' -> 'B', 'PROC-EXEC' -> 'PROC')."""
    if not code_id:
        return ""
    # Trennzeichen: '-' oder '_'
    for sep in ("-", "_"):
        if sep in code_id:
            return code_id.split(sep, 1)[0]
    # sonst erstes alphabetisches Praefix
    prefix = ""
    for ch in code_id:
        if ch.isalpha():
            prefix += ch
        else:
            break
    return prefix or code_id[:1]


def _code_name_from_sources(
    code_id: str,
    codebase_codes: dict | None,
    interview_codes: dict | None,
) -> str:
    """Sucht den Code-Namen zuerst im Codebook, dann im Interview-Result."""
    if codebase_codes and code_id in codebase_codes:
        info = codebase_codes[code_id]
        if isinstance(info, dict):
            return info.get("name") or info.get("code_name") or ""
        if isinstance(info, str):
            return info
    if interview_codes and code_id in interview_codes:
        info = interview_codes[code_id]
        if isinstance(info, dict):
            return info.get("name") or info.get("code_name") or ""
        if isinstance(info, str):
            return info
    return ""


# ---------------------------------------------------------------------------
# Interview-Rows
# ---------------------------------------------------------------------------


def _interview_rows_from_result(
    result: AnalysisResult,
    run_name: str,
    company: str,
    codebase_codes: dict | None,
) -> Iterable[list]:
    """Wandelt jedes CodedSegment in eine Excel-Zeile."""
    interview_codes = result.codes if isinstance(result.codes, dict) else {}
    for seg in result.segments:
        code_id = seg.code_id or ""
        code_name = seg.code_name or _code_name_from_sources(
            code_id, codebase_codes, interview_codes
        )
        haupt = seg.hauptkategorie or _hauptkategorie_from_code(code_id)
        yield [
            run_name,
            "Interview",
            company,
            "",  # Projekt
            seg.document or "",  # Datei
            seg.document or "",  # Pfad (kein zusaetzlicher Pfad bekannt)
            "",  # Seite (Interviews haben keine Seiten)
            code_id,
            haupt,
            code_name,
            _truncate(seg.text or ""),
            "",  # Begruendung — Mayring-Result hat keine
            int(seg.char_start or 0),
            int(seg.char_end or 0),
            "",  # Confidence
        ]


def _collect_interview_rows(
    ctx: RunContext,
    run_name: str,
    codebase_codes: dict | None,
) -> list[list]:
    """Sammelt Interview-Rows aus ``ctx.analysis_json`` UND aus
    Company-spezifischen ``<run>/<company>/analysis_results.json``."""
    rows: list[list] = []

    # Top-Level (Legacy / Transcripts-Pipeline)
    if ctx.analysis_json.exists():
        try:
            result = AnalysisResult.load(ctx.analysis_json)
            rows.extend(_interview_rows_from_result(result, run_name, "", codebase_codes))
        except Exception as e:
            print(f"  WARN: pivot_export: konnte {ctx.analysis_json} nicht laden: {e}")

    # Company-scoped
    if ctx.run_dir.exists():
        for company_dir in sorted(ctx.run_dir.iterdir()):
            if not company_dir.is_dir():
                continue
            if company_dir.name in (
                "annotated",
                "mapping",
                "qda",
                "evaluation",
                "prompts",
                "responses",
                ".cache",
                "_interview_sample",
                "_sample_input",
            ):
                continue
            candidate = company_dir / "analysis_results.json"
            if not candidate.exists():
                continue
            try:
                result = AnalysisResult.load(candidate)
                rows.extend(
                    _interview_rows_from_result(
                        result,
                        run_name,
                        company_dir.name,
                        codebase_codes,
                    )
                )
            except Exception as e:
                print(f"  WARN: pivot_export: konnte {candidate} nicht laden: {e}")

    return rows


# ---------------------------------------------------------------------------
# PDF-Rows
# ---------------------------------------------------------------------------


def _block_text_lookup(extraction: dict | None) -> dict[str, tuple[str, int, int]]:
    """Erzeugt ``{block_id: (text, char_start, char_end)}`` aus Extraktionsdaten."""
    if not extraction:
        return {}

    lookup: dict[str, tuple[str, int, int]] = {}
    offset = 0
    first = True
    for page in extraction.get("pages", []) or []:
        for block in page.get("blocks", []) or []:
            if block.get("type") != "text":
                continue
            text = block.get("text") or ""
            if not text.strip():
                continue
            if not first:
                offset += 2  # "\n\n"
            first = False
            bid = block.get("id") or ""
            char_start = offset
            char_end = offset + len(text)
            if bid:
                lookup[bid] = (text, char_start, char_end)
            offset = char_end
    return lookup


def _collect_pdf_rows(
    ctx: RunContext,
    run_name: str,
    codebase_codes: dict | None,
) -> list[list]:
    """Liest alle Codings + zugehoerige PDF-Metadaten aus der Pipeline-DB."""
    rows: list[list] = []
    db = ctx.db

    try:
        pdfs = db.get_all_pdfs()
    except Exception as e:
        print(f"  WARN: pivot_export: pdf_documents nicht lesbar: {e}")
        return rows

    company_names: dict[int, str] = {}
    try:
        conn = db._get_conn()
        company_names = {
            row["id"]: row["name"] for row in conn.execute("SELECT id, name FROM companies")
        }
    except Exception:
        company_names = {}

    for pdf in pdfs:
        pdf_id = pdf["id"]
        try:
            codings = db.get_codings_for_pdf(pdf_id)
        except Exception as e:
            print(f"  WARN: pivot_export: codings fuer pdf={pdf_id} nicht lesbar: {e}")
            continue
        if not codings:
            continue

        # Extraktion (fuer Block-Text + Char-Position) — best-effort
        extraction = None
        try:
            extraction = db.load_extraction(pdf_id)
        except Exception:
            extraction = None
        block_lookup = _block_text_lookup(extraction)

        company_name = company_names.get(pdf.get("company_id") or 0, "")

        project = pdf.get("project") or ""
        filename = pdf.get("filename") or ""
        rel_path = pdf.get("relative_path") or ""
        confidence = pdf.get("confidence") or ""

        for c in codings:
            block_id = c.get("block_id") or ""
            page = c.get("page") or 0
            begruendung = c.get("begruendung") or c.get("description") or ""
            text, char_start, char_end = block_lookup.get(block_id, ("", 0, 0))

            for code_id in c.get("codes", []):
                if not code_id:
                    continue
                code_name = _code_name_from_sources(code_id, codebase_codes, None)
                haupt = _hauptkategorie_from_code(code_id)
                rows.append(
                    [
                        run_name,
                        "Dokument",
                        company_name,
                        project,
                        filename,
                        rel_path,
                        int(page or 0),
                        code_id,
                        haupt,
                        code_name,
                        _truncate(text),
                        _truncate(begruendung, 500),
                        int(char_start or 0),
                        int(char_end or 0),
                        confidence if confidence != "" else "",
                    ]
                )

    return rows


# ---------------------------------------------------------------------------
# Excel-Schreiber
# ---------------------------------------------------------------------------


def _write_excel(rows: list[list], output_path: Path) -> None:
    """Schreibt die Rows als wide-format Sheet ``Codierungen``."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Codierungen"

    # Header
    ws.append(COLUMNS)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    for col in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        ws.append(row)

    ws.freeze_panes = "A2"

    # Auto-AutoFilter ueber Header + Daten
    ws.auto_filter.ref = ws.dimensions

    # Spaltenbreiten: einfache Heuristik (max bis 60)
    widths: dict[int, int] = {i: len(h) + 2 for i, h in enumerate(COLUMNS, 1)}
    sample_limit = min(len(rows), 200)
    for r in rows[:sample_limit]:
        for i, val in enumerate(r, 1):
            if val is None:
                continue
            length = len(str(val))
            if length > widths.get(i, 0):
                widths[i] = min(length + 2, 60)
    for i, w in widths.items():
        col_letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[col_letter].width = max(8, w)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def build_pivot_excel(
    ctx: RunContext,
    output_path: Path,
    codebase_codes: dict | None = None,
) -> int:
    """Erzeugt ``pivot_results.xlsx`` und gibt die Anzahl Datenzeilen zurueck.

    Wenn weder Interview- noch PDF-Codings vorliegen, wird KEINE Datei
    geschrieben und ``0`` zurueckgegeben.
    """
    run_name = ctx.run_dir.name
    interview_rows = _collect_interview_rows(ctx, run_name, codebase_codes)
    pdf_rows = _collect_pdf_rows(ctx, run_name, codebase_codes)

    rows = interview_rows + pdf_rows
    if not rows:
        return 0

    _write_excel(rows, Path(output_path))
    print(
        f"  Pivot-Export: {len(rows)} Codierung(en) "
        f"({len(interview_rows)} Interview, {len(pdf_rows)} Dokument) "
        f"-> {output_path}"
    )
    return len(rows)
