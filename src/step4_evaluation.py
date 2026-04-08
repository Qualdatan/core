"""Schritt 4: Auswertungs-Excel mit 3 Sheets generieren."""

from collections import Counter, defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .models import AnalysisResult


def _apply_header_style(ws, row, max_col):
    """Wendet Header-Styling an."""
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, horizontal="center")


def _apply_data_border(ws, row, max_col):
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for col in range(1, max_col + 1):
        ws.cell(row=row, column=col).border = thin_border


def generate_evaluation(result: AnalysisResult, output_path=None):
    """Generiert auswertung.xlsx mit 3 Sheets."""
    if output_path is None:
        raise ValueError("output_path muss angegeben werden")
    wb = Workbook()

    # Dokument-IDs erstellen (I-01, I-02, ...)
    doc_names = sorted(result.documents.keys())
    doc_ids = {name: f"I-{i+1:02d}" for i, name in enumerate(doc_names)}
    n_interviews = len(doc_names)

    # Segmente pro Kategorie und pro Dokument zählen
    cat_counts = Counter()       # cat -> gesamt
    cat_doc_counts = defaultdict(Counter)  # cat -> {doc -> count}
    code_counts = Counter()      # code_id -> gesamt
    code_doc_presence = defaultdict(set)   # code_id -> {doc1, doc2, ...}

    for seg in result.segments:
        cat = seg.hauptkategorie
        cat_counts[cat] += 1
        cat_doc_counts[cat][seg.document] += 1
        code_counts[seg.code_id] += 1
        code_doc_presence[seg.code_id].add(seg.document)

    # ============================================================
    # Sheet 1: Häufigkeitstabelle
    # ============================================================
    ws1 = wb.active
    ws1.title = "Häufigkeitstabelle"

    headers1 = [
        "Hauptkategorie", "Bezeichnung", "Anzahl Codes gesamt",
        "Ø Nennungen pro Interview", "Max. mögl. Nennungen", "Kodierquote %",
    ]
    for col, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=col, value=h)
    _apply_header_style(ws1, 1, len(headers1))

    ws1.column_dimensions["A"].width = 18
    ws1.column_dimensions["B"].width = 40
    ws1.column_dimensions["C"].width = 20
    ws1.column_dimensions["D"].width = 25
    ws1.column_dimensions["E"].width = 22
    ws1.column_dimensions["F"].width = 16

    row = 2
    total_segments = len(result.segments)
    categories = result.categories
    for cat_key in sorted(categories.keys()):
        cat_name = categories[cat_key]
        count = cat_counts.get(cat_key, 0)
        avg = round(count / n_interviews, 1) if n_interviews > 0 else 0
        # Max. mögliche = Codes in der Kategorie * Interviews
        n_codes_in_cat = sum(1 for cid, info in result.codes.items() if info["hauptkategorie"] == cat_key)
        max_possible = n_codes_in_cat * n_interviews if n_codes_in_cat > 0 else n_interviews
        quote = round((count / max_possible) * 100, 1) if max_possible > 0 else 0

        values = [cat_key, cat_name, count, avg, max_possible, f"{quote}%"]
        for col, val in enumerate(values, 1):
            ws1.cell(row=row, column=col, value=val)
        _apply_data_border(ws1, row, len(headers1))
        row += 1

    # Summenzeile
    sum_font = Font(bold=True)
    ws1.cell(row=row, column=1, value="GESAMT").font = sum_font
    ws1.cell(row=row, column=3, value=total_segments).font = sum_font
    avg_total = round(total_segments / n_interviews, 1) if n_interviews > 0 else 0
    ws1.cell(row=row, column=4, value=avg_total).font = sum_font
    _apply_data_border(ws1, row, len(headers1))

    # ============================================================
    # Sheet 2: Top10_Einzelcodes
    # ============================================================
    ws2 = wb.create_sheet("Top10_Einzelcodes")

    # Top 10 Codes nach Häufigkeit
    top10 = code_counts.most_common(10)

    headers2 = ["Rang", "Code-ID", "Beschreibung"]
    for doc_name in doc_names:
        headers2.append(doc_ids[doc_name])
    headers2.append("Summe")

    for col, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=h)
    _apply_header_style(ws2, 1, len(headers2))

    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 35

    row = 2
    for rang, (code_id, count) in enumerate(top10, 1):
        info = result.codes.get(code_id, {})
        ws2.cell(row=row, column=1, value=rang)
        ws2.cell(row=row, column=2, value=code_id)
        ws2.cell(row=row, column=3, value=info.get("name", ""))

        for i, doc_name in enumerate(doc_names):
            col = 4 + i
            if doc_name in code_doc_presence[code_id]:
                ws2.cell(row=row, column=col, value="✓")
                ws2.cell(row=row, column=col).alignment = Alignment(horizontal="center")
            else:
                ws2.cell(row=row, column=col, value="–")
                ws2.cell(row=row, column=col).alignment = Alignment(horizontal="center")

        ws2.cell(row=row, column=len(headers2), value=count)
        _apply_data_border(ws2, row, len(headers2))
        row += 1

    # ============================================================
    # Sheet 3: Kernergebnisse
    # ============================================================
    ws3 = wb.create_sheet("Kernergebnisse")

    headers3 = ["Nr.", "Befund", "Erläuterung"]
    for col, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=col, value=h)
    _apply_header_style(ws3, 1, len(headers3))

    ws3.column_dimensions["A"].width = 8
    ws3.column_dimensions["B"].width = 45
    ws3.column_dimensions["C"].width = 70

    wrap = Alignment(wrap_text=True, vertical="top")

    row = 2
    for i, ke in enumerate(result.kernergebnisse, 1):
        nr = ke.get("nr", i)
        ws3.cell(row=row, column=1, value=nr)
        cell_b = ws3.cell(row=row, column=2, value=ke.get("befund", ""))
        cell_b.alignment = wrap
        cell_c = ws3.cell(row=row, column=3, value=ke.get("erlaeuterung", ""))
        cell_c.alignment = wrap
        _apply_data_border(ws3, row, len(headers3))
        row += 1

    wb.save(output_path)
    print(f"  Auswertung gespeichert: {output_path}")
    print(f"  3 Sheets: Häufigkeitstabelle, Top10_Einzelcodes, Kernergebnisse")
