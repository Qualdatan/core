"""Schritt 2: Codebook als Excel generieren."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .models import AnalysisResult


def generate_codebook(result: AnalysisResult, output_path=None):
    """Generiert codebook.xlsx mit Code-ID, Name, Kategorie, Definition, Ankerbeispiel, Abgrenzung."""
    if output_path is None:
        raise ValueError("output_path muss angegeben werden")
    wb = Workbook()
    ws = wb.active
    ws.title = "Codebook"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    cat_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    cat_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    wrap = Alignment(wrap_text=True, vertical="top")

    # Header
    headers = [
        "Code-ID", "Code-Name", "Hauptkategorie",
        "Kodierdefinition", "Ankerbeispiel", "Abgrenzungsregel",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap
        cell.border = thin_border

    # Spaltenbreiten
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 45
    ws.column_dimensions["F"].width = 40

    # Codes nach Hauptkategorie sortiert einfügen
    row = 2
    codes_by_cat = {}
    for code_id, info in result.codes.items():
        cat = info["hauptkategorie"]
        codes_by_cat.setdefault(cat, []).append((code_id, info))

    for cat_key in sorted(codes_by_cat.keys()):
        # Kategorie-Überschrift
        cat_name = result.categories.get(cat_key, cat_key)
        cell = ws.cell(row=row, column=1, value=f"Kategorie {cat_key}")
        cell.font = cat_font
        cell.fill = cat_fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = cat_fill
            ws.cell(row=row, column=col).border = thin_border
        cell.value = f"Kategorie {cat_key}: {cat_name}"
        row += 1

        # Codes
        for code_id, info in sorted(codes_by_cat[cat_key]):
            values = [
                code_id,
                info["name"],
                f"{cat_key}: {cat_name}",
                info.get("kodierdefinition", ""),
                info.get("ankerbeispiel", ""),
                info.get("abgrenzungsregel", ""),
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.alignment = wrap
                cell.border = thin_border
            row += 1

    wb.save(output_path)
    print(f"  Codebook gespeichert: {output_path}")
    print(f"  {len(result.codes)} Codes in {len(codes_by_cat)} Kategorien")
